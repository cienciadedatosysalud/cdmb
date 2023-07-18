import datetime
import glob
import hashlib
import json
import os
import re
import shutil
import tempfile
import time
import uuid
from io import StringIO
from random import choices

import duckdb  # pip install duckdb==0.8.1
import graphviz
import pandas as pd
from faker import Faker
from fastapi import UploadFile
from openpyxl.reader.excel import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook import Workbook
from rocrate.model.person import Person
from rocrate.rocrate import ROCrate

from .typing.CommonDataModelTyping import JoinOptions
from .cohort.Cohort import Cohort
from .cohort.Crosswalks import Crosswalks
from .entities.Catalog import Catalog
from .entities.Entity import Entity
from .entities.Rule import DummyRule
from .entities.Variable import Variable
from .project.Author import Author
from .project.Metadata import Metadata
from .relationships.Relationship import Relationship
from .templates import Utils
from dateutil.parser import parse

def is_float(param: str):
    try:
        float(param)
        return True
    except ValueError:
        return False


def is_int(param: str):
    try:
        int(param)
        return True
    except ValueError:
        return False

def is_date(string, fuzzy=False):
    try:
        parse(string, fuzzy=fuzzy)
        return True
    except ValueError:
        return False

class CommonDataModel:
    def __init__(self,
                 metadata_definition: Metadata,
                 cohort_definition: Cohort,
                 entities: list[Entity],
                 relationships: list[Relationship] = None
                 ):
        """Initialize a Common Data Model object.

    A Common Data Model is a standardized way of representing data from different sources and domains. It consists of metadata, cohort, entities and relationships.

    Args:
        metadata_definition (Metadata): Metadata structure.
        cohort_definition (Cohort): Cohort structure.
        entities (list[Entity]): List of Entity structures.
        relationships (list[Relationship]): List of Relationship structures. Defaults to None.
    """

        self._created_dt = datetime.datetime.today().strftime('%Y-%m-%d %H:%M:%S')
        self._uuid = str(uuid.uuid4())
        self._out_dir = ""
        self._metadata: Metadata = metadata_definition
        self._cohort: Cohort = cohort_definition
        self._entities: list[Entity] = entities
        self._relationships: list[Relationship] = relationships if relationships is not None else []
        self.folder_code = self._created_dt = datetime.datetime.today().strftime('%Y%m%d%H%M%S') + '_cdmb'

    @property
    def created_dt(self):
        return self._created_dt

    @property
    def metadata(self):
        return self._metadata

    @metadata.setter
    def metadata(self, metadata_: Metadata):
        self._metadata = metadata_

    @property
    def cohort(self):
        return self._cohort

    @cohort.setter
    def cohort(self, cohort_: Cohort):
        self._cohort = cohort_

    @property
    def entities(self):
        return self._entities

    @entities.setter
    def entities(self, entities_: [Entity]):
        self._entities = entities_

    def add_entity(self, entity: Entity):
        self._entities.append(entity)

    def pop_entity(self):
        return self._entities.pop()

    @property
    def relationships(self):
        return self._relationships

    @relationships.setter
    def relationships(self, relationships_: [Relationship]):
        self._relationships = relationships_

    #############################
    # Private auxiliary methods #
    #############################

    def __getRootPath(self):
        return os.path.join(self._out_dir, self.folder_code)

    def __getDocsPath(self):
        return os.path.join(self._out_dir, self.folder_code, 'docs', "CDM")

    def __getDocsPathCDM(self):
        return os.path.join(self._out_dir, self.folder_code, 'docs', "CDM", "common_datamodel.xlsx")

    def __getValidationRulesPath(self, entity_name):
        return os.path.join(self._out_dir, self.folder_code, 'docs', "CDM", "entities", entity_name,
                            "validation-rules")

    def __getCatalogPath(self, entity_name):
        return os.path.join(self._out_dir, self.folder_code, 'docs', "CDM", "entities", entity_name, "catalogs")

    def __getSyntheticPath(self):
        return os.path.join(self._out_dir, self.folder_code, 'docs', "CDM", "synthetic-data")

    def __getInputPath(self):
        return os.path.join(self._out_dir, self.folder_code, 'inputs')

    def __getSrcPath(self):
        return os.path.join(self._out_dir, self.folder_code, 'src')

    def __create_work_structure(self):
        structure = [
            os.path.join(self._out_dir, self.folder_code),
            os.path.join(self._out_dir, self.folder_code, 'inputs'),
            os.path.join(self._out_dir, self.folder_code, 'outputs', 'logs'),
            os.path.join(self._out_dir, self.folder_code, 'src', "dqa-scripts"),
            os.path.join(self._out_dir, self.folder_code, 'src', "validation-scripts"),
            os.path.join(self._out_dir, self.folder_code, 'src', "check_load-scripts", "inputs"),
            os.path.join(self._out_dir, self.folder_code, 'src', "analysis-scripts"),
            os.path.join(self._out_dir, self.folder_code, 'docs', "CDM", "synthetic-data")
        ]
        for entity in self._entities:
            structure.append(
                os.path.join(self._out_dir, self.folder_code, 'docs', "CDM", "entities", entity.name, "catalogs")
            )
            structure.append(
                os.path.join(self._out_dir, self.folder_code, 'docs', "CDM", "entities", entity.name,
                             "validation-rules")
            )
        for path in structure:
            os.makedirs(path)

    def save_project(self, out_dir: str = "."):
        """
        Args:
         out_dir (String): Output directory

        """
        self._out_dir = out_dir
        self.__create_work_structure()
        self.__write_metadata()
        self.__write_cohort()
        page = self.__write_crosswalk()
        self.__write_entities(page)
        self.__write_rules()
        self.__write_configuration_file()
        self.__write_catalogs()
        self.__generate_synthetic()
        self.__save_er()
        Utils.generate_documentation(self.__getRootPath())
        self.__create_rog()
        self.__generate_md5()

    def save_zipped_project(self, out_dir: str = "."):
        """

        Args:
         out_dir (String): Output directory.

        """
        with tempfile.TemporaryDirectory(dir=out_dir) as tmpdirname:
            self.save_project(str(tmpdirname))
            time.sleep(1)
            uuid_zip = str(uuid.uuid4())
            shutil.make_archive(os.path.join(out_dir, self.folder_code), 'zip', root_dir=str(tmpdirname),
                                base_dir=self.folder_code)
            time.sleep(1)
            return uuid_zip

    def generate_json_structure(self) -> dict:
        response = {
            "created_dt": self._created_dt,
            "uuid": self._uuid,
            'metadata': self.metadata.get_structure(),
            'cohort': self.cohort.get_structure(),
            'entities': [e.get_structure() for e in self.entities],
            'relationships': [r.get_structure() for r in self.relationships]
        }

        return response

    # Operaciones para editar Excel
    def __write_metadata(self):
        path_ = self.__getDocsPathCDM()
        workbook = Workbook()
        worksheet = workbook.create_sheet("Metadata", 0)
        worksheet.cell(row=1, column=1, value='Project')
        worksheet.cell(row=1, column=2, value=self.metadata.project)
        worksheet.cell(row=2, column=1, value='Project URL')
        worksheet.cell(row=2, column=2, value=self.metadata.url_project)
        worksheet.cell(row=3, column=1, value='Work package')
        worksheet.cell(row=3, column=2, value=self.metadata.work_package)
        worksheet.cell(row=4, column=1, value='Use case')
        worksheet.cell(row=4, column=2, value=self.metadata.use_case)
        worksheet.cell(row=5, column=1, value='Document')
        worksheet.cell(row=5, column=2, value=self.metadata.document)
        worksheet.cell(row=6, column=1, value='Version (SEM)')
        worksheet.cell(row=6, column=2, value=self.metadata.version_sem)
        worksheet.cell(row=7, column=1, value='Authors')
        worksheet.cell(row=7, column=2, value='Name')
        worksheet.cell(row=7, column=3, value='Affiliation')
        worksheet.cell(row=7, column=4, value='ORCID')
        row = 7
        for author_ in self.metadata.authors:
            row = row + 1
            worksheet.cell(row=row, column=2, value=author_.name)
            worksheet.cell(row=row, column=3, value=author_.affiliation)
            worksheet.cell(row=row, column=4, value=author_.id)
        row = row + 1
        worksheet.cell(row=row, column=1, value='Keywords')
        worksheet.cell(row=row, column=2, value=','.join(self.metadata.keywords))
        worksheet.cell(row=row + 1, column=1, value='Description')
        worksheet.cell(row=row + 1, column=2, value=self.metadata.description)
        worksheet.cell(row=row + 2, column=1, value='Notes')
        worksheet.cell(row=row + 2, column=2, value=self.metadata.notes)

        workbook.save(filename=path_)

    def __write_cohort(self):
        path_ = self.__getDocsPathCDM()
        workbook = load_workbook(path_)
        worksheet = workbook.create_sheet("Cohort definition", 1)
        worksheet.cell(row=1, column=1, value='Cohort name')
        worksheet.cell(row=1, column=2, value=self.cohort.name)
        worksheet.cell(row=2, column=1, value='Cohort description')
        worksheet.cell(row=2, column=2, value=self.cohort.description)
        worksheet.cell(row=3, column=1, value='Inclusion criteria')
        worksheet.cell(row=3, column=2, value=self.cohort.inclusion_criteria)
        worksheet.cell(row=4, column=1, value='Exclusion criteria')
        worksheet.cell(row=4, column=2, value=self.cohort.exclusion_criteria)
        worksheet.cell(row=5, column=1, value='Beginning of study period')
        worksheet.cell(row=5, column=2, value=self.cohort.beggining_study_period)
        worksheet.cell(row=6, column=1, value='End of study period')
        worksheet.cell(row=6, column=2, value=self.cohort.end_study_period)

        worksheet.cell(row=8, column=1, value='Cohort definition catalog (inclusion)')
        worksheet.cell(row=9, column=1, value='Nature')
        nature = self.cohort.cohort_definition_inclusion.nature if self.cohort.cohort_definition_inclusion is not None else None
        column_name = self.cohort.cohort_definition_inclusion.column_name if self.cohort.cohort_definition_inclusion is not None else None
        worksheet.cell(row=9, column=2, value=nature)
        worksheet.cell(row=10, column=1, value='Column name')
        worksheet.cell(row=10, column=2, value=column_name)

        worksheet.cell(row=12, column=1, value='Cohort definition catalog (exclusion)')
        worksheet.cell(row=13, column=1, value='Nature')
        nature = self.cohort.cohort_definition_exclusion.nature if self.cohort.cohort_definition_exclusion is not None else None
        column_name = self.cohort.cohort_definition_exclusion.column_name if self.cohort.cohort_definition_exclusion is not None else None
        worksheet.cell(row=13, column=2, value=nature)
        worksheet.cell(row=14, column=1, value='Column name')
        worksheet.cell(row=14, column=2, value=column_name)
        worksheet.cell(row=15, column=1,
                       value='The crosswalks data can be found in the following document sheets (If there are)')
        workbook.save(filename=path_)

    def __write_crosswalk(self):
        page = 1
        if self.cohort.cohort_definition_inclusion is not None:
            df = self.cohort.cohort_definition_inclusion.data
            path_ = self.__getDocsPathCDM()
            path_csv = os.path.join(self.__getDocsPath(), 'cohort_definition_inclusion.csv')
            workbook = load_workbook(path_)
            if df is not None:
                page = page + 1
                worksheet = workbook.create_sheet("cohort_definition_inclusion", page)
                for r in dataframe_to_rows(df, index=False, header=True):
                    worksheet.append(r)
                df.to_csv(path_csv, index=False)
            workbook.save(filename=path_)

        if self.cohort.cohort_definition_exclusion is not None:
            df = self.cohort.cohort_definition_exclusion.data
            path_ = self.__getDocsPathCDM()
            path_csv = os.path.join(self.__getDocsPath(), 'cohort_definition_exclusion.csv')
            workbook = load_workbook(path_)
            if df is not None:
                page = page + 1
                worksheet = workbook.create_sheet("cohort_definition_exclusion", page)
                for r in dataframe_to_rows(df, index=False, header=True):
                    worksheet.append(r)
                df.to_csv(path_csv, index=False)
            workbook.save(filename=path_)
        return page + 1

    def __write_entities(self, page):

        path_ = self.__getDocsPathCDM()
        # page = 4
        for entity in self.entities:
            workbook = load_workbook(path_)
            worksheet = workbook.create_sheet(entity.name, page)
            label_ = [var.label for var in entity.variables]
            description_ = [var.description for var in entity.variables]
            standard_classification_ = [var.standard_classification for var in entity.variables]
            format_ = [var.format for var in entity.variables]
            type_ = [var.type for var in entity.variables]
            units_ = [var.units for var in entity.variables]
            requirement_level_ = [var.requirement_level for var in entity.variables]
            characteristic_ = [var.characteristic for var in entity.variables]
            catalog_bl_ = [var.catalog_bl for var in entity.variables]
            transformations_from_origin_ = [var.transformations_from_origin for var in entity.variables]
            possible_data_source_ = [var.possible_data_source for var in entity.variables]
            observations_comments_ = [var.observations_comments for var in entity.variables]
            examples_ = [var.examples for var in entity.variables]
            df = pd.DataFrame({'label': label_,
                               'description': description_,
                               'standard_classification': standard_classification_,
                               'format': format_,
                               'type_': type_,
                               'units': units_,
                               'requirement_level': requirement_level_,
                               'characteristic': characteristic_,
                               'catalog_bl': catalog_bl_,
                               'transformations_from_origin': transformations_from_origin_,
                               'possible_data_source': possible_data_source_,
                               'observations_comments': observations_comments_,
                               'examples': examples_})
            for r in dataframe_to_rows(df, index=False, header=True):
                worksheet.append(r)
            workbook.save(filename=path_)
            page = page + 1

    def __write_rules(self):
        for entity in self.entities:
            path_ = os.path.join(self.__getValidationRulesPath(entity.name),
                                 'rules_set_' + str(entity.name).replace(' ', '_') + '.json')
            rules_dict = entity.rules.get_rules_dict()
            with open(path_, 'x') as f:
                json_object = json.dumps(rules_dict, ensure_ascii=False, indent=4)
                f.write(json_object)

    def __write_catalogs(self):
        for entity in self.entities:
            for variable in entity.variables:
                if variable.catalog_bl:
                    if variable.catalog is not None:
                        df = variable.catalog.data
                        path_csv = os.path.join(self.__getCatalogPath(entity.name), variable.catalog.filename)
                        path_ = self.__getDocsPathCDM()
                        workbook = load_workbook(path_)
                        clean_entity_name = str(entity.name).replace(' ', '_')
                        worksheet = workbook.create_sheet(clean_entity_name + '_' + variable.label, -1)
                        if df is not None:
                            for r in dataframe_to_rows(df, index=False, header=True):
                                worksheet.append(r)
                            df.to_csv(path_csv, index=False)
                        workbook.save(filename=path_)

    def __write_configuration_file(self):
        path_ = os.path.join(self.__getDocsPath(),
                             'cdmb_config.json')
        with open(path_, 'x') as f:
            json_object = json.dumps(self.generate_json_structure(), ensure_ascii=False, indent=4)
            f.write(json_object)

    def __save_er(self):
        e = graphviz.Graph(name='ER', engine='neato', directory=self.__getDocsPath())
        e.attr('node', shape='box')
        for entity in self.entities:
            e.node(entity.name, fontsize="10pt")
            e.attr('node', shape='ellipse')
            for variable in entity.variables:
                if len(variable.label) > 15:
                    e.node(entity.name + '_' + variable.label, label=variable.label, fontsize="8pt")
                    e.edge(entity.name + '_' + variable.label, entity.name, len='2.1')
                else:
                    e.node(entity.name + '_' + variable.label, label=variable.label, fontsize="8pt")
                    e.edge(entity.name + '_' + variable.label, entity.name, len='1.5')
            e.attr('node', shape='box')

        e.attr('node', shape='diamond', style='filled', color='lightgrey')
        for relation in self.relationships:
            if relation.right_entity is not None and relation.right_column is not None:
                e.node(relation.left_entity.name + '-' + relation.right_entity.name, label=relation.join_type,
                       fontsize='10')
                e.edge(relation.left_entity.name, relation.left_entity.name + '-' + relation.right_entity.name,
                       label=relation.left_column.label,
                       len='2.5')
                e.edge(relation.right_entity.name, relation.left_entity.name + '-' + relation.right_entity.name,
                       label=relation.right_column.label,
                       len='2.5')
        e.attr(label=r'\n' + self.cohort.name)
        e.attr(fontsize='15')
        e.render(directory=self.__getDocsPath(), format='png')

    def __check_if_entity_in_relation(self, entity):
        response = None
        for relationship in self._relationships:
            if relationship.left_entity.name == entity.name:
                response = relationship.left_column
                break
            elif relationship.right_entity is not None and relationship.right_entity == entity.name:
                response = relationship.right_column
                break
        return response

    def __generate_synthetic(self, num_registries: int = 1000):
        database_ = os.path.join(self.__getInputPath(), 'data.duckdb')
        con = duckdb.connect(database=database_, read_only=False)
        fake = Faker()
        for entity in self.entities:
            # If related entity reset seed to match.
            Faker.seed(1234)
            rules_set = entity.get_rules_synthetic()
            catalog_by_variable, catalog_by_filename = entity.get_catalogs()
            path_ = os.path.join(self.__getSyntheticPath(), 'synthetic_' + str(entity.name).replace(' ', '_') + '.csv')
            relation_variable = self.__check_if_entity_in_relation(entity)
            if relation_variable is None:
                variables_list = entity.variables
            else:
                variables_list = entity.get_variables_in_order(relation_variable)

            df = pd.DataFrame()
            variables_to_avoid = []
            for variable in variables_list:
                label = variable.label
                data_ = {}
                if label not in variables_to_avoid:
                    format_ = variable.format
                    # check if label in Catalogs
                    if label in catalog_by_variable:
                        catalog_set = catalog_by_filename[catalog_by_variable[label]]
                        variable_list = catalog_set["variable_list"]
                        if format_ != "Boolean":
                            # If the variable is Boolean, the catalog only determines its definition, not its values.
                            data = catalog_set["data"]
                            data = data[variable_list].sample(
                                n=num_registries,
                                replace=True)
                            data.reset_index(drop=True, inplace=True)
                            df = pd.concat([df, data], axis=1)
                            variables_to_avoid.extend(variable_list)
                        else:
                            data_[label] = [fake.pybool() for _ in range(num_registries)]
                            df = pd.concat([df, pd.DataFrame(data_)], axis=1)
                            variables_to_avoid.extend(variable_list)

                    elif label in rules_set and format_ != "Boolean":
                        ## check if label in rules
                        rule = rules_set[label][0]
                        if rule['category'] == "restricted":
                            data_[label] = choices(rule["values"],
                                                   k=num_registries)
                        else:
                            # String Categorical variable that takes numerical/date variables.
                            if format_ == "String" and (is_int(rule['min_value']) or is_int(rule['max_value'])):
                                format_ = "Integer"
                            elif format_ == "String" and (is_float(rule['min_value']) or is_float(rule['max_value'])):
                                format_ = "Double"
                            elif format_ == "String" and (is_date(rule['min_value']) or is_date(rule['max_value'])):
                                format_ = "Datetime"

                            if 'interval_comparison' in rule['category'] and rule['negative'] is True:
                                # not between X and Y
                                left_side_registries = int(num_registries / 2)
                                right_side_registries = num_registries - left_side_registries
                                if format_ == "Date":
                                    data_[label] = [fake.date_between_dates(date_end=rule['min_value'])
                                                    for _ in range(left_side_registries)] + \
                                                   [fake.date_between_dates(date_start=rule['max_value'])
                                                    for _ in range(right_side_registries)]
                                elif format_ == "Datetime":
                                    data_[label] = [fake.date_time_between(end_date=rule['min_value'])
                                                    for _ in range(left_side_registries)] + \
                                                   [fake.date_time_between(start_date=rule['max_value'])
                                                    for _ in range(right_side_registries)]
                                elif format_ == "Integer":
                                    data_[label] = [fake.pyint(max_value=rule['max_value']) for _ in
                                                    range(left_side_registries)] + [
                                                       fake.pyint(min_value=rule['max_value'])
                                                       for _ in
                                                       range(right_side_registries)]

                                elif format_ == "Double":
                                    data_[label] = [fake.pyfloat(positive=True, left_digits=5, right_digits=2,
                                                                 max_value=rule['min_value'])
                                                    for _
                                                    in range(left_side_registries)] + \
                                                   [fake.pyfloat(left_digits=5, right_digits=2,
                                                                 min_value=rule['max_value'])
                                                    for _
                                                    in range(right_side_registries)]

                            else:
                                if format_ == "Date":
                                    if rule['min_value'] is None:
                                        data_[label] = [
                                            fake.date_between_dates(date_start=datetime.date(1900, 1, 1),
                                                                    date_end=rule['max_value']) for _ in
                                            range(num_registries)]
                                    elif rule['max_value'] is None:
                                        data_[label] = [
                                            fake.date_between_dates(date_start=rule['min_value'],
                                                                    date_end=datetime.date.today()) for _ in
                                            range(num_registries)]
                                    else:
                                        data_[label] = [
                                            fake.date_between_dates(date_start=rule['min_value'],
                                                                    date_end=rule['max_value']) for _ in
                                            range(num_registries)]
                                elif format_ == "Datetime":
                                    if rule['min_value'] is None:
                                        data_[label] = [
                                            fake.date_time_between(end_date=rule['max_value'])
                                            for _ in range(num_registries)]
                                    elif rule['max_value'] is None:
                                        data_[label] = [
                                            fake.date_time_between(start_date=rule['min_value'])
                                            for _ in range(num_registries)]
                                    else:
                                        data_[label] = [
                                            fake.date_time_between(start_date=rule['min_value'],
                                                                   end_date=rule['max_value'])
                                            for _ in range(num_registries)]
                                elif format_ == "Integer":
                                    if rule['max_value'] is not None and rule['min_value'] is not None:
                                        data_[label] = [
                                            fake.pyint(min_value=rule['min_value'], max_value=rule['max_value'])
                                            for _ in
                                            range(num_registries)]
                                    if rule['max_value'] is None and rule['min_value'] is not None:
                                        data_[label] = [fake.pyint(min_value=rule['min_value']) for _ in
                                                        range(num_registries)]
                                    if rule['max_value'] is not None and rule['min_value'] is None:
                                        data_[label] = [fake.pyint(max_value=rule['max_value']) for _ in
                                                        range(num_registries)]
                                elif format_ == "Double":
                                    data_[label] = [
                                        fake.pyfloat(left_digits=5, right_digits=2, min_value=rule['min_value'],
                                                     max_value=rule['max_value']) for _
                                        in range(num_registries)]
                        df = pd.concat([df, pd.DataFrame(data_)], axis=1)
                    else:
                        # No rules, no catalog
                        if format_ == "String":
                            data_['' + label + ''] = [fake.pystr() for _ in range(num_registries)]
                        if format_ == "Boolean":
                            data_[label] = [fake.pybool() for _ in range(num_registries)]
                        if format_ == "Date":
                            data_[label] = [fake.date_between() for _ in range(num_registries)]
                        if format_ == "Datetime":
                            data_[label] = [fake.date_time_between() for _ in range(num_registries)]
                        if format_ == "Integer":
                            data_[label] = [fake.pyint() for _ in range(num_registries)]
                        if format_ == "Double":
                            data_[label] = [fake.pyfloat(positive=True, left_digits=5, right_digits=2) for _ in
                                            range(num_registries)]
                        df = pd.concat([df, pd.DataFrame(data_)], axis=1)

            tree_dependencies = entity.get_tree_structure_rules(variables_list)

            def change_dependency_values(child, parent):
                if len(child.children) == 0 and parent is None:
                    pass
                elif len(child.children) == 0 and parent is not None:
                    # Cambiar valor con el padre
                    if child.variable.format == "Date":
                        max_value = max(df[parent.label])
                        df[child.variable.label] = df[parent.label].apply(
                            lambda x: fake.date_between(start_date=x, end_date=max_value))
                    elif child.variable.format == "Datetime":
                        max_value = max(df[parent.label])
                        df[child.variable.label] = df[parent.label].apply(
                            lambda x: fake.date_time_between(start_date=x, end_date=max_value))
                    elif child.variable.format == "Integer":
                        max_value = max(df[parent.label])
                        df[child.variable.label] = df[parent.label].apply(
                            lambda x: fake.pyint(min_value=x, max_value=max_value))
                    elif child.variable.format == "Double":
                        max_value = max(df[parent.label])
                        df[child.variable.label] = df[parent.label].apply(
                            lambda x: fake.pyfloat(min_value=x, max_value=max_value))
                else:
                    for child_ in child.children:
                        change_dependency_values(child_, child.variable)

            for child in tree_dependencies.children:
                change_dependency_values(child, None)

            if str(entity.name).strip() == "":
                raise ValueError("The entity name cannot be an empty string.")

            df.to_csv(path_, index=False)

            con.sql("CREATE TABLE " + entity.name + " AS SELECT * FROM df")

    def __create_rog(self):
        crate = ROCrate()
        authors = []
        orcid_prefix = "https://orcid.org/"
        crate.name = str(self.metadata.project or '')
        crate.description = str(self.metadata.description or '')
        crate.keywords = ', '.join(self.metadata.keywords)
        crate.license = str(self.metadata.license or '')
        crate.creator = "Common Data Model Builder"
        for author in self.metadata.authors:
            orcid_id_ = author.id
            name_ = author.name
            affiliation_ = author.affiliation
            if orcid_id_ is not None:
                exp1 = re.search("[0-9]{4}-[0-9]{4}-[0-9]{4}-[0-9]{4}", orcid_id_)
                if exp1 is not None:
                    orcid_id_ = orcid_prefix + orcid_id_
                    authors.append(crate.add(Person(crate, orcid_id_, properties={
                        "name": name_,
                        "affiliation": affiliation_
                    })))
                else:
                    authors.append(crate.add(Person(crate, properties={
                        "name": name_,
                        "affiliation": affiliation_
                    })))
            authors.append(crate.add(Person(crate, properties={
                "name": name_,
                "affiliation": affiliation_
            })))

        # Create a list of digital objects that are generated.
        path_readme = os.path.join(self.__getRootPath(), 'README.md')
        readme_file = crate.add_file(path_readme, dest_path=path_readme.replace(self.__getRootPath() + '/', ''),
                                     properties={
                                         "@type": "File",
                                         "name": "README",
                                         "encodingFormat": "text/markdown"
                                     })
        readme_file["author"] = authors

        path_license = os.path.join(self.__getRootPath(), 'LICENSE.md')
        license_ = crate.add_file(path_license, dest_path=path_license.replace(self.__getRootPath() + '/', ''),
                                  properties={
                                      "@type": "File",
                                      "name": "License",
                                      "encodingFormat": "text/markdown"
                                  })
        license_["author"] = authors

        path_man_container = os.path.join(self.__getRootPath(), 'man_container_deployment.md')
        manual_file = crate.add_file(path_man_container,
                                     dest_path=path_man_container.replace(self.__getRootPath() + '/', ''), properties={
                "@type": "File",
                "name": "Manual on how to deploy the project in a docker container",
                "encodingFormat": "text/markdown"
            })
        manual_file["author"] = authors

        path_git_container = os.path.join(self.__getRootPath(), '.gitignore')
        git_file = crate.add_file(path_git_container,
                                  dest_path=path_git_container.replace(self.__getRootPath() + '/', ''), properties={
                "@type": "File",
                "name": ".gitignore file",
                "encodingFormat": "text/plain"
            })
        git_file["author"] = authors

        # input
        path_duckdb = os.path.join(self.__getInputPath(), 'data.duckdb')
        duckdb_file = crate.add_file(path_duckdb, dest_path=path_duckdb.replace(self.__getRootPath() + '/', ''),
                                     properties={
                                         "@type": "File",
                                         "name": "SQL OLAP auxiliary database",
                                         "encodingFormat": "application/octet-stream"
                                     })
        duckdb_file["author"] = authors

        # Scripting
        path_dqa = os.path.join(self.__getSrcPath(), 'dqa-scripts', 'dqa.py')
        dqa_file = crate.add_file(path_dqa, dest_path=path_dqa.replace(self.__getRootPath() + '/', ''), properties={
            "@type": ["File", "SoftwareSourceCode"],
            "name": "Data Quality Assessment script",
            "encodingFormat": "text/plain"
        })
        dqa_file["author"] = authors

        path_validator = os.path.join(self.__getSrcPath(), 'validation-scripts', 'validator.py')
        validator_file = crate.add_file(path_validator,
                                        dest_path=path_validator.replace(self.__getRootPath() + '/', ''), properties={
                "@type": ["File", "SoftwareSourceCode"],
                "name": "Rules validator file",
                "encodingFormat": "text/plain"
            })
        validator_file["author"] = authors

        path_validator_report = os.path.join(self.__getSrcPath(), 'validation-scripts', 'validator_report.qmd')
        validator_report = crate.add_file(path_validator_report,
                                          dest_path=path_validator_report.replace(self.__getRootPath() + '/', ''),
                                          properties={
                                              "@type": ["File", "SoftwareSourceCode"],
                                              "name": "Rules validator report",
                                              "encodingFormat": "text/plain"
                                          })
        validator_report["author"] = authors

        path_validator_metadata = os.path.join(self.__getSrcPath(), 'validation-scripts', '_quarto.yml')
        validator_report_metadata = crate.add_file(path_validator_metadata,
                                                   dest_path=path_validator_metadata.replace(self.__getRootPath() + '/',
                                                                                             ''), properties={
                "@type": ["File", "SoftwareSourceCode"],
                "name": "Rules validator report metadata",
                "encodingFormat": "text/yaml"
            })
        validator_report_metadata["author"] = authors

        path_templates_r = os.path.join(self.__getSrcPath(), "analysis-scripts", 'r_report_template.qmd')
        template_quarto_file = crate.add_file(path_templates_r,
                                              dest_path=path_templates_r.replace(self.__getRootPath() + '/', ''),
                                              properties={
                                                  "@type": ["File", "SoftwareSourceCode"],
                                                  "name": "Quarto analysis script template",
                                                  "encodingFormat": "text/plain"
                                              })
        template_quarto_file["author"] = authors

        path_analysis_metadata = os.path.join(self.__getSrcPath(), 'analysis-scripts', '_quarto.yml')
        analysis__report_metadata = crate.add_file(path_analysis_metadata,
                                                   dest_path=path_analysis_metadata.replace(self.__getRootPath() + '/',
                                                                                            ''), properties={
                "@type": ["File", "SoftwareSourceCode"],
                "name": "Analysis script report metadata",
                "encodingFormat": "text/yaml"
            })
        analysis__report_metadata["author"] = authors

        # Docs
        path_config = os.path.join(self.__getDocsPath(), "cdmb_config.json")
        config_file = crate.add_file(path_config, dest_path=path_config.replace(self.__getRootPath() + '/', ''),
                                     properties={
                                         "@type": "File",
                                         "name": "CDM builder and project configuration file",
                                         "encodingFormat": "application/json"
                                     })
        config_file["author"] = authors

        path_cdm = self.__getDocsPathCDM()

        datamodel = crate.add_file(path_cdm, dest_path=path_cdm.replace(self.__getRootPath() + '/', ''), properties={
            "@type": "File",
            "name": "Common Data Model spreadsheet",
            "encodingFormat": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        })
        datamodel["author"] = authors

        path_crosswalks = os.path.join(self.__getDocsPath(), "cohort_definition_inclusion.csv")
        if os.path.exists(path_crosswalks):
            crosswalks = crate.add_file(path_crosswalks,
                                        dest_path=path_crosswalks.replace(self.__getRootPath() + '/', ''), properties={
                    "@type": "File",
                    "name": "Crosswalks for the cohort definition inclusion",
                    "encodingFormat": "text/csv"
                })
            crosswalks["author"] = authors

        path_crosswalks = os.path.join(self.__getDocsPath(), "cohort_definition_exclusion.csv")
        if os.path.exists(path_crosswalks):
            crosswalks = crate.add_file(path_crosswalks,
                                        dest_path=path_crosswalks.replace(self.__getRootPath() + '/', ''), properties={
                    "@type": "File",
                    "name": "Crosswalks for the cohort definition exclusion",
                    "encodingFormat": "text/csv"
                })
            crosswalks["author"] = authors

        er_png_path = os.path.join(self.__getDocsPath(), "ER.gv.png")
        er_diagram_png = crate.add_file(er_png_path, dest_path=er_png_path.replace(self.__getRootPath() + '/', ''),
                                        properties={
                                            "@type": "File",
                                            "name": "Entity Relationship Diagram PNG",
                                            "encodingFormat": "image/png"
                                        })
        er_diagram_png["author"] = authors

        er_gv_path = os.path.join(self.__getDocsPath(), "ER.gv")
        er_diagram_gv = crate.add_file(er_gv_path, dest_path=er_gv_path.replace(self.__getRootPath() + '/', ''),
                                       properties={
                                           "@type": "File",
                                           "name": "Entity Relationship Diagram GV",
                                           "encodingFormat": "text/plain"
                                       })
        er_diagram_gv["author"] = authors

        # Entities files
        for entity in self.entities:
            # Synthetic data
            synthetic_path = os.path.join(self.__getSyntheticPath(),
                                          'synthetic_' + str(entity.name).replace(' ', '_') + '.csv')
            synthetic_data = crate.add_file(synthetic_path,
                                            dest_path=synthetic_path.replace(self.__getRootPath() + '/', ''),
                                            properties={
                                                "@type": "File",
                                                "name": "Synthetic data for entity: " + entity.name,
                                                "encodingFormat": "text/csv"
                                            })
            synthetic_data["author"] = authors
            # catalogs
            for variable in entity.variables:
                if variable.catalog_bl:
                    if variable.catalog is not None:
                        catalog_path = os.path.join(self.__getCatalogPath(entity.name), variable.catalog.filename)
                        catalog_data = crate.add_file(catalog_path,
                                                      dest_path=catalog_path.replace(self.__getRootPath() + '/', ''),
                                                      properties={
                                                          "@type": "File",
                                                          "name": "Catalog file for " + entity.name + "-" + variable.label,
                                                          "encodingFormat": "text/csv"
                                                      })
                        catalog_data["author"] = authors

            rules_path = os.path.join(self.__getValidationRulesPath(entity.name),
                                      'rules_set_' + str(entity.name).replace(' ', '_') + '.json')
            rules_data = crate.add_file(rules_path, dest_path=rules_path.replace(self.__getRootPath() + '/', ''),
                                        properties={
                                            "@type": "File",
                                            "name": "Rule set for entity: " + entity.name,
                                            "encodingFormat": "application/json"
                                        })
            rules_data["author"] = authors
        crate.write_crate(self.__getRootPath())

    def __generate_md5(self):
        outdir = self._out_dir + '/'
        hash_file = {"files": []}
        files_to_has = glob.glob(outdir + '**', recursive=True)
        for filename in files_to_has:
            if os.path.isfile(filename):
                md5_hash = hashlib.md5()
                with open(filename, "rb") as f:
                    # Read and update hash in chunks of 4K
                    for byte_block in iter(lambda: f.read(4096), b""):
                        md5_hash.update(byte_block)
                    hash_file['files'].append(
                        {
                            "filename": filename.replace(outdir, ''),
                            "hash": md5_hash.hexdigest()
                        }
                    )
        with open(os.path.join(self.__getDocsPath(), 'hashed_files_list.json'), 'w') as outfile:
            json_object = json.dumps(hash_file, indent=4)
            outfile.write(json_object)

    @staticmethod
    def load_previous_configuration(configuration: dict):
        """

        Args:
            configuration (dict): json with a previous configuration
        Returns:
            metadata (Metadata): f
            cohort (Cohort): c
            entities (list[Entity]): e
            relationships (list[Relationship]): r
        """

        # Create Metadata
        if 'metadata' in configuration:
            main_keys = {"project", "funder", "url_project", "work_package", "use_case", "document", "version_sem",
                         "keywords", "description", "notes", "spatial_coverage", "license"}
            core_info = {k: v for k, v in configuration['metadata'].items() if k in main_keys}
            author_info = [Author(**v) for v in configuration['metadata']['authors']]
            core_info["authors"] = author_info
            metadata = Metadata(**core_info)
        else:
            raise ValueError('The "metadata" field does not exist in the configuration file and is a mandatory field.')

        # Create Cohort
        if 'cohort' in configuration:
            main_keys = {"name", "description", "inclusion_criteria", "exclusion_criteria"}
            dates_keys = {"beggining_study_period", "end_study_period"}
            core_info = {k: v for k, v in configuration['cohort'].items() if k in main_keys}
            date_ifo = {k: pd.to_datetime(v) for k, v in configuration['cohort'].items() if k in dates_keys}
            core_info = core_info | date_ifo
            cohort = Cohort(**core_info)
        else:
            raise ValueError('The "cohort" field does not exist in the configuration file and is a mandatory field.')
        # Create Entities
        entities_list = []
        entities_catalog = {}
        if 'entities' in configuration:
            for entity in configuration['entities']:
                main_keys = {"name", "time_varying"}
                core_info = {k: v for k, v in entity.items() if k in main_keys}
                if 'variables' not in entity:
                    raise ValueError(
                        'The "entities["variables"]" field does not exist in the configuration file and is a mandatory field.')

                def remove_catalog(x):
                    x.pop('catalog')
                    return x

                variables_info = map(remove_catalog, entity['variables'])
                variables_info_ = []
                for variable in variables_info:
                    variable_clean = {k: v for k, v in variable.items() if v is not None}
                    variables_info_.append(Variable(**variable_clean))
                # variables_info = [Variable(**v) for v in variables_info]

                core_info["variables"] = variables_info_

                entity_ = Entity(**core_info)

                if 'rules' not in entity:
                    raise ValueError(
                        'The "entities["rules"]" field does not exist in the configuration file and is a mandatory field.')

                # Rules
                for rule in entity['rules']:
                    if 'expression' not in rule:
                        raise ValueError("Error in the structure of the rule. The key 'expression' must be present.")
                    if 'name' not in rule:
                        raise ValueError("Error in the structure of the rule. The key 'name' must be present.")
                    if 'description' not in rule:
                        raise ValueError("Error in the structure of the rule. The key 'description' must be present.")
                    expression_ = rule['expression']
                    rule_ = None
                    try:
                        rule_ = entity_.create_rule_from_expression(expression_, rule['name'], rule['description'])
                    except Exception as e_:
                        rule_ = DummyRule(expression_, rule['name'], rule['description'])
                        pass
                    if rule_ is not None:
                        entity_.rules.append_rule(rule_)
                # append
                entities_catalog[entity_.name] = entity_
                entities_list.append(entity_)
        else:
            raise ValueError('The "entities" field does not exist in the configuration file and is a mandatory field.')
        # Create Relationships
        if 'relationships' in configuration:
            relationships_list: list[Relationship] = []
            for relationship in configuration['relationships']:
                if 'left_entity' not in relationship:
                    raise ValueError("'left_entity' is not declared in the relationship")
                if 'right_entity' not in relationship:
                    raise ValueError("'right_entity' is not declared in the relationship")
                if 'join_type' not in relationship:
                    raise ValueError("'join_type' is not declared in the relationship")
                if 'left_column' not in relationship:
                    raise ValueError("'left_column' is not declared in the relationship")
                if 'right_column' not in relationship:
                    raise ValueError("'right_column' is not declared in the relationship")
                left_entity_str = relationship["left_entity"]
                if left_entity_str not in entities_catalog:
                    raise ValueError("The left entity of the relationship does not match any of the declared entities.")
                left_entity = entities_catalog[left_entity_str]

                right_entity_str = relationship["right_entity"]
                if right_entity_str not in entities_catalog:
                    raise ValueError("The right entity of the relationship does not match any of the declared entities.")
                right_entity = entities_catalog[right_entity_str]

                join_type = relationship["join_type"]
                if join_type not in JoinOptions.__args__:
                    raise ValueError('join_type must be in {literal_values}'.format(literal_values=JoinOptions.__args__))

                left_column = left_entity.get_variable_by_label(relationship["left_column"])
                if left_column is None:
                    raise ValueError('"left_column" variable has to be part of its entity ("left_entity")')

                right_column = right_entity.get_variable_by_label(relationship["right_column"])
                if right_column is None:
                    raise ValueError('"right_column" variable has to be part of its entity ("right_entity")')

                relationships_list.append(Relationship(left_entity, right_entity, left_column, right_column, join_type))
        else:
            raise ValueError(
                'The "relationships" field does not exist in the configuration file and is a mandatory field.')
        return metadata, cohort, entities_list, relationships_list

    @staticmethod
    def load_previous_configuration_from_web(configuration: dict, files: list[UploadFile]):
        """

        Args:
            configuration (dict): json with a previous configuration
            configuration (list[UploadFile]): files from cdmb ui
        Returns:
            metadata (Metadata): f
            cohort (Cohort): c
            entities (list[Entity]): e
            relationships (list[Relationship]): r
        """

        def infer_separator(stringio_):
            firstline = stringio_.readline().rstrip()
            stringio_.seek(0)
            separators = re.sub('"*[a-zA-ZÀ-ÿñÑ0-9_-]*"*', '', firstline)
            if len(separators) > 0:
                return separators[0]
            else:
                # Return random separator, exception will be thrown on header reading
                return '|'

        def create_variable(x, _files):
            catalog = None
            if 'catalog' in x:
                catalog = x.pop('catalog')
                if catalog is not None and ('filename' not in catalog or 'column_name' not in catalog):
                    catalog = None
            variable_clean = {k: v for k, v in x.items() if v is not None}
            variable = Variable(**variable_clean)
            if variable.catalog_bl and catalog is not None:
                filename_temp = catalog['filename']
                columnname_ = catalog['column_name']
                for file_ in _files:
                    if file_.file.closed is False and filename_temp == file_.filename:
                        contents_ = file_.file.read()
                        s_ = str(contents_, 'utf-8')
                        data_ = StringIO(s_)
                        separator_ = infer_separator(data_)
                        df_catalog = pd.read_csv(data_, sep=separator_)
                        data_.close()
                        file_.file.close()
                        variable.catalog = Catalog(df_catalog, columnname_, filename_temp)
                        break
            return variable

        # Create Metadata
        if 'metadata' in configuration:
            main_keys = {"project", "funder", "url_project", "work_package", "use_case", "document", "version_sem",
                         "keywords", "description", "notes", "spatial_coverage", "license"}
            core_info = {k: v for k, v in configuration['metadata'].items() if k in main_keys}
            author_info = [Author(**v) for v in configuration['metadata']['authors']]
            core_info["authors"] = author_info
            metadata = Metadata(**core_info)
        else:
            raise ValueError('The "metadata" field does not exist in the configuration file and is a mandatory field.')

        # Create Cohort
        if 'cohort' in configuration:
            main_keys = {"name", "description", "inclusion_criteria", "exclusion_criteria"}
            dates_keys = {"beggining_study_period", "end_study_period"}
            core_info = {k: v for k, v in configuration['cohort'].items() if k in main_keys}
            date_ifo = {k: pd.to_datetime(v) for k, v in configuration['cohort'].items() if k in dates_keys}
            core_info = core_info | date_ifo
            cohort = Cohort(**core_info)

            filename_ = configuration['cohort']['cohort_definition_inclusion']['filename']
            if filename_ != '':
                for file in files:
                    if file.file.closed is False and filename_ == file.filename:
                        contents = file.file.read()
                        s = str(contents, 'utf-8')
                        data = StringIO(s)
                        separator = infer_separator(data)
                        df_crosswalks = pd.read_csv(data, sep=separator)
                        data.close()
                        file.file.close()
                        cohort.cohort_definition_inclusion = Crosswalks(df_crosswalks,
                                                                        configuration['cohort'][
                                                                            'cohort_definition_inclusion'][
                                                                            'column_name'],
                                                                        configuration['cohort'][
                                                                            'cohort_definition_inclusion'][
                                                                            'nature'],
                                                                        filename_)
                        break
            filename_ = configuration['cohort']['cohort_definition_exclusion']['filename']
            if filename_ != '':
                for file in files:
                    if file.file.closed is False and filename_ == file.filename:
                        contents = file.file.read()
                        s = str(contents, 'utf-8')
                        data = StringIO(s)
                        separator = infer_separator(data)
                        df_crosswalks = pd.read_csv(data, sep=separator)
                        data.close()
                        file.file.close()
                        cohort.cohort_definition_exclusion = Crosswalks(df_crosswalks,
                                                                        configuration['cohort'][
                                                                            'cohort_definition_exclusion'][
                                                                            'column_name'],
                                                                        configuration['cohort'][
                                                                            'cohort_definition_exclusion'][
                                                                            'nature'],
                                                                        filename_)
        else:
            raise ValueError('The "cohort" field does not exist in the configuration file and is a mandatory field.')

        # Create Entities
        entities_list = []
        entities_catalog = {}
        if 'entities' in configuration:
            for entity in configuration['entities']:
                main_keys = {"name", "time_varying"}
                core_info = {k: v for k, v in entity.items() if k in main_keys}
                if 'variables' not in entity:
                    raise ValueError(
                        'The "entities["variables"]" field does not exist in the configuration file and is a mandatory field.')

                variables_info = [create_variable(v, files) for v in entity['variables'] if v]

                core_info["variables"] = variables_info

                entity_ = Entity(**core_info)

                if 'rules' not in entity:
                    raise ValueError(
                        'The "entities["rules"]" field does not exist in the configuration file and is a mandatory field.')

                # Rules
                for rule in entity['rules']:
                    if 'expression' not in rule:
                        raise ValueError("Error in the structure of the rule. The key 'expression' must be present.")
                    if 'name' not in rule:
                        raise ValueError("Error in the structure of the rule. The key 'name' must be present.")
                    if 'description' not in rule:
                        raise ValueError("Error in the structure of the rule. The key 'description' must be present.")
                    expression_ = rule['expression']
                    rule_ = None
                    try:
                        rule_ = entity_.create_rule_from_expression(expression_, rule['name'], rule['description'])
                    except Exception as e_:
                        rule_ = DummyRule(expression_, rule['name'], rule['description'])
                        pass
                    if rule_ is not None:
                        entity_.rules.append_rule(rule_)
                # append
                entities_catalog[entity_.name] = entity_
                entities_list.append(entity_)
        else:
            raise ValueError('The "entities" field does not exist in the configuration file and is a mandatory field.')

        # Create Relationships
        if 'relationships' in configuration:
            relationships_list: list[Relationship] = []
            for relationship in configuration['relationships']:
                if 'left_entity' not in relationship:
                    raise ValueError("'left_entity' is not declared in the relationship")
                if 'right_entity' not in relationship:
                    raise ValueError("'right_entity' is not declared in the relationship")
                if 'join_type' not in relationship:
                    raise ValueError("'join_type' is not declared in the relationship")
                if 'left_column' not in relationship:
                    raise ValueError("'left_column' is not declared in the relationship")
                if 'right_column' not in relationship:
                    raise ValueError("'right_column' is not declared in the relationship")
                left_entity_str = relationship["left_entity"]
                if left_entity_str not in entities_catalog:
                    raise ValueError("The left entity of the relationship does not match any of the declared entities.")
                left_entity = entities_catalog[left_entity_str]

                right_entity_str = relationship["right_entity"]
                if right_entity_str not in entities_catalog:
                    raise ValueError(
                        "The right entity of the relationship does not match any of the declared entities.")
                right_entity = entities_catalog[right_entity_str]

                join_type = relationship["join_type"]
                if join_type not in JoinOptions.__args__:
                    raise ValueError(
                        'join_type must be in {literal_values}'.format(literal_values=JoinOptions.__args__))

                left_column = left_entity.get_variable_by_label(relationship["left_column"])
                if left_column is None:
                    raise ValueError('"left_column" variable has to be part of its entity ("left_entity")')

                right_column = right_entity.get_variable_by_label(relationship["right_column"])
                if right_column is None:
                    raise ValueError('"right_column" variable has to be part of its entity ("right_entity")')
                relationships_list.append(Relationship(left_entity, right_entity, left_column, right_column, join_type))
        else:
            raise ValueError(
                'The "relationships" field does not exist in the configuration file and is a mandatory field.')
        return metadata, cohort, entities_list, relationships_list

    @staticmethod
    def get_er(entities: list[Entity], relationships: list[Relationship]):
        with tempfile.TemporaryDirectory() as tmpdirname:
            e = graphviz.Graph(name='ER', engine='neato', directory=tmpdirname)
            e.attr('node', shape='box')
            for entity in entities:
                e.node(entity.name, fontsize="10pt")
                e.attr('node', shape='ellipse')
                for variable in entity.variables:
                    if len(variable.label) > 15:
                        e.node(entity.name + '_' + variable.label, label=variable.label, fontsize="8pt")
                        e.edge(entity.name + '_' + variable.label, entity.name, len='2.1')
                    else:
                        e.node(entity.name + '_' + variable.label, label=variable.label, fontsize="8pt")
                        e.edge(entity.name + '_' + variable.label, entity.name, len='1.5')
                e.attr('node', shape='box')

            e.attr('node', shape='diamond', style='filled', color='lightgrey')
            for relation in relationships:
                if relation.right_entity is not None and relation.right_column is not None:
                    e.node(relation.left_entity.name + '-' + relation.right_entity.name, label=relation.join_type,
                           fontsize='10')
                    e.edge(relation.left_entity.name, relation.left_entity.name + '-' + relation.right_entity.name,
                           label=relation.left_column.label,
                           len='2.5')
                    e.edge(relation.right_entity.name, relation.left_entity.name + '-' + relation.right_entity.name,
                           label=relation.right_column.label,
                           len='2.5')
            e.attr(fontsize='15')
            return e.pipe(format='png')
